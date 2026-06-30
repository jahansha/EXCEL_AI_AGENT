"""
Generate a holdings report from MerrylAccount.xlsx (workbook sheet 'Merrill').

Rules (from CLAUDE.md):
- Open the Excel file READ ONLY.
- Process every row that has the letter 'f' in column A (the "current row").
- Ticker: from the current row move to column B and look UP (row above, then the
  row above that, ...) until a non-empty value is found. That value is the ticker.
- Capture count from column D and unit price from column E of the current row.
- Write the report to the output/ folder.

New instructions #1:
- For each data row also display the current price (column F), total cost
  (column G), current value (column H), plus computed Profit/Loss and the
  percentage of change.
- Total cost and current value are taken from the 'f' (aggregate) row.
- The current price column on the 'f' row holds a weight fraction, so the
  current price is read from the lot row(s) above (same upward walk as the
  ticker), since all lots of a ticker share the same market price.

New instructions #2:
- Include the sheet row number for each ticker (the 'f'/current row processed).
- Order the report by ascending Profit / Loss (largest losses first).

New instructions #3:
- Move the 'Sheet Row' column one position left (ahead of 'Ticker').
- Append a summary report: total cost, total amount, and percentage for the
  losing positions (P/L < 0) and, separately, the winning positions (P/L > 0).

New instructions #4:
- If the program is passed the parameter 'OBT' the report is Ordered By Ticker
  (alphabetical). Otherwise it keeps the default ascending Profit / Loss order.

New instructions #5:
- Add a 'Rating' column to the LEFT of the 'Ticker' column. Its value is taken
  from column B of the current ('f') row itself.
- The output filename now starts with 'profit_' (was 'profi_').

New instructions #6:
- Write the report to the 'output' directory located in the PARENT directory of
  this script's directory (i.e. ../output relative to this file).
"""

import os
import sys
from datetime import datetime

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)

EXCEL_PATH = r"C:\Users\johna\Documents\MerrylAccount.xlsx"
SHEET_NAME = "Merrill"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# New instructions #6: the 'output' directory lives in the PARENT directory of
# this script's directory.
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(PARENT_DIR, "output")


def is_blank(value):
    """Treat None and whitespace-only cells as blank."""
    return value is None or str(value).strip() == ""


def find_up(rows, current_index, col_idx):
    """Walk upward from the current row in a column until a non-empty value."""
    for r in range(current_index - 1, -1, -1):
        if col_idx < len(rows[r]):
            val = rows[r][col_idx]
            if not is_blank(val):
                return val
    return None


def find_current_price(rows, current_index, count, current_value,
                       col_a_idx=0, col_f_idx=5):
    """Current price from column F of the lot rows above the 'f' row.

    The 'f' row and any intermediate sub-total row carry a *weight fraction*
    (always between 0 and 1) in column F instead of a price, while genuine
    per-share prices are >= 1. So we walk upward and return the first column-F
    value >= 1, stopping if we cross into the previous ticker's 'f' row. If no
    price-like cell is found we fall back to current_value / count.
    """
    for r in range(current_index - 1, -1, -1):
        a = rows[r][col_a_idx] if col_a_idx < len(rows[r]) else None
        if a is not None and str(a).strip().lower() == "f":
            break  # crossed into the previous ticker block without a price
        f = to_float(rows[r][col_f_idx]) if col_f_idx < len(rows[r]) else None
        if f is not None and f >= 1.0:
            return f
    cnt = to_float(count)
    val = to_float(current_value)
    if cnt not in (None, 0) and val is not None:
        return val / cnt  # derived fallback
    return None


def to_float(value):
    """Best-effort numeric conversion; returns None when not numeric."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main():
    # Load entire sheet into memory (read-only).
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    # Materialize rows as a list of tuples (0-based column indexing).
    rows = [row for row in ws.iter_rows(values_only=True)]

    COL_A, COL_B, COL_D, COL_E = 0, 1, 3, 4
    COL_F, COL_G, COL_H = 5, 6, 7  # Current Price, Total Cost, Current Value

    def cell(row, idx):
        return row[idx] if len(row) > idx else None

    records = []
    for i, row in enumerate(rows):
        a = row[COL_A] if len(row) > COL_A else None
        if a is not None and str(a).strip().lower() == "f":
            ticker = find_up(rows, i, COL_B)
            ticker = str(ticker).strip() if ticker is not None else "(unknown)"

            # Rating: column B of the current ('f') row itself (not the upward
            # walk used for the ticker).
            rating = cell(row, COL_B)
            rating = str(rating).strip() if not is_blank(rating) else "-"

            count = cell(row, COL_D)
            unit_price = cell(row, COL_E)
            total_cost = to_float(cell(row, COL_G))
            current_value = to_float(cell(row, COL_H))
            # Current price lives on the lot rows above (the 'f' row and any
            # intermediate sub-total row hold a weight fraction in column F).
            current_price = find_current_price(
                rows, i, count, current_value, COL_A, COL_F)

            profit_loss = None
            pct_change = None
            if total_cost is not None and current_value is not None:
                profit_loss = current_value - total_cost
                if total_cost != 0:
                    pct_change = profit_loss / total_cost * 100.0

            records.append({
                "row": i + 1,            # 1-based for display
                "rating": rating,
                "ticker": ticker,
                "count": count,
                "unit_price": unit_price,
                "current_price": current_price,
                "total_cost": total_cost,
                "current_value": current_value,
                "profit_loss": profit_loss,
                "pct_change": pct_change,
            })

    wb.close()

    # Ordering: 'OBT' (Order By Ticker) -> alphabetical by ticker; otherwise the
    # default ascending Profit / Loss (largest losses first), rows without a
    # P/L value placed last.
    order_by_ticker = any(a.strip().upper() == "OBT" for a in sys.argv[1:])
    if order_by_ticker:
        records.sort(key=lambda r: r["ticker"].upper())
        order_desc = "ordered by ticker (A-Z)"
    else:
        records.sort(key=lambda r: (r["profit_loss"] is None,
                                    r["profit_loss"] if r["profit_loss"]
                                    is not None else 0.0))
        order_desc = "sorted ascending by Profit / Loss"

    # ---- Build the PDF ----
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now = datetime.now()
    # Requested pattern: profit_loss_YYYY_MM_DD:HH:MM.pdf -> ':' is illegal on
    # Windows, so colons are replaced with dashes.
    filename = f"profit_loss_{now:%Y_%m_%d-%H-%M}.pdf"
    out_path = os.path.join(OUTPUT_DIR, filename)

    doc = SimpleDocTemplate(
        out_path, pagesize=landscape(letter),
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        title="Merrill Profit / Loss Report",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=10, textColor=colors.grey)

    elements = []
    elements.append(Paragraph("Merrill Profit / Loss Report", title_style))
    elements.append(Paragraph(
        f"Source: {os.path.basename(EXCEL_PATH)} &nbsp;|&nbsp; Sheet: {SHEET_NAME}"
        f" &nbsp;|&nbsp; Generated: {now:%Y-%m-%d %H:%M:%S}", sub_style))
    elements.append(Paragraph(
        f"Holdings processed (rows flagged 'f' in column A): {len(records)}"
        f" &nbsp;|&nbsp; {order_desc}", sub_style))
    elements.append(Spacer(1, 0.25 * inch))

    def fmt_qty(v):
        f = to_float(v)
        if f is None:
            return "-" if v in (None, "") else str(v)
        return f"{f:,.0f}" if f == int(f) else f"{f:,.4f}"

    def fmt_money(v):
        f = to_float(v)
        return "-" if f is None else f"${f:,.2f}"

    def fmt_pct(v):
        f = to_float(v)
        return "-" if f is None else f"{f:+.2f}%"

    def fmt_signed_money(v):
        f = to_float(v)
        return "-" if f is None else f"${f:+,.2f}"

    # 'Sheet Row' moved one column left (now ahead of 'Ticker'); 'Rating'
    # inserted to the left of 'Ticker'.
    headers = ["#", "Sheet Row", "Rating", "Ticker", "Count", "Unit Price",
               "Current Price", "Total Cost", "Current Value",
               "Profit / Loss", "% Change"]
    data = [headers]

    tot_cost = tot_value = 0.0
    # Loss / gain breakdown accumulators for the summary section.
    loss_cost = loss_amt = 0.0
    gain_cost = gain_amt = 0.0
    loss_n = gain_n = 0
    pl_row_indices = []  # body rows that carry a Profit/Loss value (for coloring)
    for n, rec in enumerate(records, start=1):
        data.append([
            str(n),
            str(rec["row"]),
            rec["rating"],
            rec["ticker"],
            fmt_qty(rec["count"]),
            fmt_money(rec["unit_price"]),
            fmt_money(rec["current_price"]),
            fmt_money(rec["total_cost"]),
            fmt_money(rec["current_value"]),
            fmt_signed_money(rec["profit_loss"]),
            fmt_pct(rec["pct_change"]),
        ])
        if rec["total_cost"] is not None:
            tot_cost += rec["total_cost"]
        if rec["current_value"] is not None:
            tot_value += rec["current_value"]
        pl, tc = rec["profit_loss"], rec["total_cost"]
        if pl is not None and tc is not None:
            if pl < 0:
                loss_cost += tc
                loss_amt += pl
                loss_n += 1
            elif pl > 0:
                gain_cost += tc
                gain_amt += pl
                gain_n += 1
        pl_row_indices.append((len(data) - 1, rec["profit_loss"]))

    # Totals row
    tot_pl = tot_value - tot_cost
    tot_pct = (tot_pl / tot_cost * 100.0) if tot_cost else None
    data.append([
        "", "", "", "TOTAL", "", "", "",
        fmt_money(tot_cost), fmt_money(tot_value),
        fmt_signed_money(tot_pl), fmt_pct(tot_pct),
    ])

    table = Table(
        data,
        colWidths=[0.3*inch, 0.65*inch, 0.6*inch, 0.8*inch, 0.85*inch,
                   0.95*inch, 1.0*inch, 1.15*inch, 1.15*inch, 1.15*inch,
                   0.8*inch],
        repeatRows=1,
    )
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (3, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#eef2f9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b0b7c3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        # Totals row emphasis
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#d6dce8")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1f3864")),
    ])

    # Color Profit/Loss and % Change green (gain) or red (loss) per row.
    green = colors.HexColor("#0a7d28")
    red = colors.HexColor("#c0241c")
    for ridx, pl in pl_row_indices:
        if pl is None:
            continue
        color = green if pl >= 0 else red
        style.add("TEXTCOLOR", (9, ridx), (10, ridx), color)
    style.add("TEXTCOLOR", (9, -1), (10, -1), green if tot_pl >= 0 else red)

    table.setStyle(style)
    elements.append(table)

    # ---- Summary report: losing vs winning positions ----
    loss_pct = (loss_amt / loss_cost * 100.0) if loss_cost else None
    gain_pct = (gain_amt / gain_cost * 100.0) if gain_cost else None
    net_cost = loss_cost + gain_cost
    net_amt = loss_amt + gain_amt
    net_pct = (net_amt / net_cost * 100.0) if net_cost else None

    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Summary", styles["Heading2"]))

    summary_data = [
        ["Category", "Positions", "Total Cost", "Total Profit / Loss", "% of Cost"],
        ["Losing positions (P/L < 0)", str(loss_n), fmt_money(loss_cost),
         fmt_signed_money(loss_amt), fmt_pct(loss_pct)],
        ["Winning positions (P/L > 0)", str(gain_n), fmt_money(gain_cost),
         fmt_signed_money(gain_amt), fmt_pct(gain_pct)],
        ["Net (all positions)", str(loss_n + gain_n), fmt_money(net_cost),
         fmt_signed_money(net_amt), fmt_pct(net_pct)],
    ]
    summary = Table(
        summary_data,
        colWidths=[2.2*inch, 0.9*inch, 1.4*inch, 1.6*inch, 1.0*inch],
        hAlign="LEFT",
    )
    summary_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b0b7c3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TEXTCOLOR", (3, 1), (4, 1), red),
        ("TEXTCOLOR", (3, 2), (4, 2), green),
        ("LINEABOVE", (0, 3), (-1, 3), 1, colors.HexColor("#1f3864")),
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#eef2f9")),
        ("TEXTCOLOR", (3, 3), (4, 3), green if net_amt >= 0 else red),
    ])
    summary.setStyle(summary_style)
    elements.append(summary)

    doc.build(elements)

    # Console summary
    print(f"Processed {len(records)} 'f' rows ({order_desc}).")
    print(f"{'Rating':<8}{'Ticker':<8}{'Row':>5}{'Count':>10}{'CurPrice':>12}"
          f"{'TotCost':>14}{'CurValue':>14}{'P/L':>14}{'%Chg':>9}")
    for rec in records:
        print(f"{rec['rating']:<8}{rec['ticker']:<8}{rec['row']:>5}"
              f"{fmt_qty(rec['count']):>10}"
              f"{fmt_money(rec['current_price']):>12}"
              f"{fmt_money(rec['total_cost']):>14}"
              f"{fmt_money(rec['current_value']):>14}"
              f"{fmt_signed_money(rec['profit_loss']):>14}"
              f"{fmt_pct(rec['pct_change']):>9}")
    print(f"{'TOTAL':<16}{'':>5}{'':>10}{'':>12}{fmt_money(tot_cost):>14}"
          f"{fmt_money(tot_value):>14}{fmt_signed_money(tot_pl):>14}"
          f"{fmt_pct(tot_pct):>9}")

    print("\nSummary")
    print(f"{'Category':<28}{'Pos':>5}{'Total Cost':>16}{'Total P/L':>16}{'% Cost':>9}")
    print(f"{'Losing positions (P/L<0)':<28}{loss_n:>5}{fmt_money(loss_cost):>16}"
          f"{fmt_signed_money(loss_amt):>16}{fmt_pct(loss_pct):>9}")
    print(f"{'Winning positions (P/L>0)':<28}{gain_n:>5}{fmt_money(gain_cost):>16}"
          f"{fmt_signed_money(gain_amt):>16}{fmt_pct(gain_pct):>9}")
    print(f"{'Net (all positions)':<28}{loss_n+gain_n:>5}{fmt_money(net_cost):>16}"
          f"{fmt_signed_money(net_amt):>16}{fmt_pct(net_pct):>9}")
    print(f"\nPDF written to: {out_path}")


if __name__ == "__main__":
    main()
