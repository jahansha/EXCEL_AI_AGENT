"""
Generate an INDIVIDUAL (per-lot, per-account) profit/loss report from
MerrylAccount.xlsx (workbook sheet 'Merrill').

Background / how the sheet is laid out
--------------------------------------
Stocks are purchased in up to three accounts: E, N and R. Each ticker forms a
"block" of rows:

  * One or more LOT rows  - the individual purchases. Column C holds the
    account (E / N / R); column B holds the ticker on the first lot of the
    block (later lots in the same block leave column B blank); columns D-H hold
    Count, Unit Price, Current Price, Total Cost and Current Value for that lot.
  * An (optional) sub-total row - column C holds a weight fraction, not an
    account.
  * An 'f' row (letter 'f' in column A) - the all-accounts AGGREGATE for the
    ticker. Here column B is a rating letter and column C is a weight fraction,
    NOT an account.

Because 19 of 34 tickers are held across two or three accounts, the 'f' row
cannot be attributed to a single account. This "Individual" report therefore
works at the LOT level: one report line per purchase lot (every row whose
column C is E, N or R), which is the only level at which a per-account
profit / loss is well defined.

What each report row shows
--------------------------
- Sheet row number (new instructions #2)
- Account (column C of the lot row)
- Ticker (column B of the lot row, or the nearest non-blank value above it)
- Count (column D), Unit Price (column E)
- Current Price (column F), Total Cost (column G), Current Value (column H)
- Profit / Loss and the percentage of change (computed, new instructions #1)

Rows are ordered ascending by Profit / Loss (largest losses first - new
instructions #2). A per-account summary is appended so profit / loss can be
read per account (E, N, R).

Output: written READ ONLY from the source workbook to the 'output' folder in
the PARENT directory of this script's directory, named
profit_loss_Individual_YYYY_MM_DD-HH-MM.pdf.
"""

import os
import sys
from datetime import datetime

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)

EXCEL_PATH = r"C:\Users\johna\Documents\MerrylAccount.xlsx"
SHEET_NAME = "Merrill"
ACCOUNTS = ("E", "N", "R")
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# 'output' folder lives in the PARENT directory of this script's directory.
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(PARENT_DIR, "output")

COL_A, COL_B, COL_C, COL_D, COL_E = 0, 1, 2, 3, 4
COL_F, COL_G, COL_H = 5, 6, 7  # Current Price, Total Cost, Current Value


def is_blank(value):
    """Treat None and whitespace-only cells as blank."""
    return value is None or str(value).strip() == ""


def to_float(value):
    """Best-effort numeric conversion; returns None when not numeric."""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cell(row, idx):
    return row[idx] if len(row) > idx else None


def find_ticker(rows, lot_index):
    """Ticker from column B of the lot row, or the nearest non-blank value
    above it (later lots in a block leave column B blank). Values may carry
    stray whitespace (e.g. 'NVDA ') so they are stripped."""
    for r in range(lot_index, -1, -1):
        val = cell(rows[r], COL_B)
        if not is_blank(val):
            return str(val).strip()
    return "(unknown)"


def main():
    # Load entire sheet into memory (read-only).
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows = [row for row in ws.iter_rows(values_only=True)]
    wb.close()

    records = []
    for i, row in enumerate(rows):
        acct = cell(row, COL_C)
        # A lot row is any row whose column C is one of the account codes.
        # This naturally excludes 'f' rows and sub-total rows (column C there
        # is a weight fraction) and blank rows (column C is empty).
        if is_blank(acct) or str(acct).strip().upper() not in ACCOUNTS:
            continue

        account = str(acct).strip().upper()
        ticker = find_ticker(rows, i)
        count = cell(row, COL_D)
        unit_price = cell(row, COL_E)
        current_price = to_float(cell(row, COL_F))
        total_cost = to_float(cell(row, COL_G))
        current_value = to_float(cell(row, COL_H))

        # Compute Profit / Loss and % change (the sheet's stored % column is
        # inconsistently scaled, so we derive both ourselves).
        profit_loss = None
        pct_change = None
        if total_cost is not None and current_value is not None:
            profit_loss = current_value - total_cost
            if total_cost != 0:
                pct_change = profit_loss / total_cost * 100.0

        records.append({
            "row": i + 1,            # 1-based sheet row
            "account": account,
            "ticker": ticker,
            "count": count,
            "unit_price": unit_price,
            "current_price": current_price,
            "total_cost": total_cost,
            "current_value": current_value,
            "profit_loss": profit_loss,
            "pct_change": pct_change,
        })

    # New instructions #4: if the parameter 'OBT' is passed, order each report
    # by ticker name (alphabetical); otherwise keep the default ascending
    # Profit / Loss order (new instructions #2 - largest losses first), with
    # rows lacking a P/L value placed last.
    order_by_ticker = any(a.strip().upper() == "OBT" for a in sys.argv[1:])
    if order_by_ticker:
        records.sort(key=lambda r: r["ticker"].upper())
        order_desc = "ordered by ticker (A-Z)"
    else:
        records.sort(key=lambda r: (r["profit_loss"] is None,
                                    r["profit_loss"] if r["profit_loss"]
                                    is not None else 0.0))
        order_desc = "sorted ascending by Profit / Loss"

    # ---- Formatting helpers ----
    def fmt_qty(v):
        f = to_float(v)
        if f is None:
            return "-" if v in (None, "") else str(v)
        return f"{f:,.0f}" if f == int(f) else f"{f:,.4f}"

    def fmt_money(v):
        f = to_float(v)
        return "-" if f is None else f"${f:,.2f}"

    def fmt_pct(v):
        f = to_float(v)
        return "-" if f is None else f"{f:+.2f}%"

    def fmt_signed_money(v):
        f = to_float(v)
        return "-" if f is None else f"${f:+,.2f}"

    # ---- Build the PDF ----
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    now = datetime.now()
    # Requested pattern: profit_loss_Individual_YYYY_MM_DD:HH:MM.pdf -> ':' is
    # illegal on Windows, so colons are replaced with dashes.
    filename = f"profit_loss_Individual_{now:%Y_%m_%d-%H-%M}.pdf"
    out_path = os.path.join(OUTPUT_DIR, filename)

    doc = SimpleDocTemplate(
        out_path, pagesize=landscape(letter),
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        title="Merrill Individual (Per-Account) Profit / Loss Report",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=10, textColor=colors.grey)

    elements = []
    elements.append(Paragraph(
        "Merrill Individual Profit / Loss Report (per Account)", title_style))
    elements.append(Paragraph(
        f"Source: {os.path.basename(EXCEL_PATH)} &nbsp;|&nbsp; Sheet: {SHEET_NAME}"
        f" &nbsp;|&nbsp; Generated: {now:%Y-%m-%d %H:%M:%S}", sub_style))
    elements.append(Paragraph(
        f"Individual purchase lots processed: {len(records)}"
        f" &nbsp;|&nbsp; grouped by account (Edge-N, Roth, Edge)"
        f" &nbsp;|&nbsp; {order_desc}", sub_style))
    elements.append(Spacer(1, 0.25 * inch))

    # New instructions #3: divide the report into three groups by account.
    #   N -> "Edge-N",  R -> "Roth",  E -> "Edge"
    # The Account column is dropped from the group tables (it is constant per
    # group and stated in each group heading instead).
    GROUP_LABELS = [("N", "Edge-N"), ("R", "Roth"), ("E", "Edge")]

    green = colors.HexColor("#0a7d28")
    red = colors.HexColor("#c0241c")

    group_headers = ["#", "Sheet Row", "Ticker", "Count", "Unit Price",
                     "Current Price", "Total Cost", "Current Value",
                     "Profit / Loss", "% Change"]
    group_col_widths = [0.35*inch, 0.7*inch, 0.85*inch, 0.9*inch, 1.0*inch,
                        1.05*inch, 1.2*inch, 1.2*inch, 1.2*inch, 0.85*inch]

    def build_group(code, label, recs):
        """Build the heading + table for one account group.

        Returns (heading, table, stats) where stats holds the group's cost,
        value, total loss (sum of negative P/L lots) and total profit (sum of
        positive P/L lots). Rows keep the chosen ordering (the input list is
        already sorted and filtering preserves order). A subtotal row is added.
        """
        data = [group_headers]
        g_cost = g_value = 0.0
        g_loss = g_profit = 0.0  # new instructions #4
        pl_row_indices = []  # body rows carrying a P/L value (for coloring)
        for n, rec in enumerate(recs, start=1):
            data.append([
                str(n),
                str(rec["row"]),
                rec["ticker"],
                fmt_qty(rec["count"]),
                fmt_money(rec["unit_price"]),
                fmt_money(rec["current_price"]),
                fmt_money(rec["total_cost"]),
                fmt_money(rec["current_value"]),
                fmt_signed_money(rec["profit_loss"]),
                fmt_pct(rec["pct_change"]),
            ])
            if rec["total_cost"] is not None:
                g_cost += rec["total_cost"]
            if rec["current_value"] is not None:
                g_value += rec["current_value"]
            pl = rec["profit_loss"]
            if pl is not None:
                if pl < 0:
                    g_loss += pl
                elif pl > 0:
                    g_profit += pl
            pl_row_indices.append((len(data) - 1, rec["profit_loss"]))

        g_pl = g_value - g_cost
        g_pct = (g_pl / g_cost * 100.0) if g_cost else None
        data.append([
            "", "", "SUBTOTAL", "", "", "",
            fmt_money(g_cost), fmt_money(g_value),
            fmt_signed_money(g_pl), fmt_pct(g_pct),
        ])

        table = Table(data, colWidths=group_col_widths, repeatRows=1)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.5),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (2, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#eef2f9")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b0b7c3")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            # Subtotal row emphasis
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#d6dce8")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1f3864")),
        ])
        # Color Profit/Loss and % Change green (gain) or red (loss) per row.
        for ridx, pl in pl_row_indices:
            if pl is None:
                continue
            style.add("TEXTCOLOR", (8, ridx), (9, ridx),
                      green if pl >= 0 else red)
        style.add("TEXTCOLOR", (8, -1), (9, -1), green if g_pl >= 0 else red)
        table.setStyle(style)

        # New instructions #4: header shows this account's Total Loss and
        # Total Profit.
        heading = Paragraph(
            f"{label} &nbsp;&mdash;&nbsp; Account {code} &nbsp;&mdash;&nbsp;"
            f" {len(recs)} lots &nbsp;&mdash;&nbsp;"
            f" <font color='#c0241c'>Total Loss: {fmt_signed_money(g_loss)}</font>"
            f" &nbsp;|&nbsp;"
            f" <font color='#0a7d28'>Total Profit: {fmt_signed_money(g_profit)}</font>",
            styles["Heading2"])
        return heading, table, {"cost": g_cost, "value": g_value,
                                "loss": g_loss, "profit": g_profit}

    tot_cost = tot_value = 0.0
    tot_loss = tot_profit = 0.0
    # Per-account accumulators for the summary section.
    acct_stats = {a: {"n": 0, "cost": 0.0, "value": 0.0,
                      "loss": 0.0, "profit": 0.0} for a in ACCOUNTS}
    for code, label in GROUP_LABELS:
        recs = [r for r in records if r["account"] == code]
        heading, table, gs = build_group(code, label, recs)
        elements.append(heading)
        elements.append(Spacer(1, 0.08 * inch))
        elements.append(table)
        elements.append(Spacer(1, 0.28 * inch))
        acct_stats[code]["n"] = len(recs)
        acct_stats[code]["cost"] = gs["cost"]
        acct_stats[code]["value"] = gs["value"]
        acct_stats[code]["loss"] = gs["loss"]
        acct_stats[code]["profit"] = gs["profit"]
        tot_cost += gs["cost"]
        tot_value += gs["value"]
        tot_loss += gs["loss"]
        tot_profit += gs["profit"]

    # Grand totals (across all accounts)
    tot_pl = tot_value - tot_cost
    tot_pct = (tot_pl / tot_cost * 100.0) if tot_cost else None

    # ---- Per-account summary ----
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph("Profit / Loss by Account", styles["Heading2"]))

    # New instructions #4: the final table carries a Total Losses column and a
    # Total Profits column for each account.
    summary_data = [["Account", "Lots", "Total Cost", "Current Value",
                     "Total Losses", "Total Profits", "Profit / Loss",
                     "% Change"]]
    for a in ACCOUNTS:
        st = acct_stats[a]
        pl = st["value"] - st["cost"]
        pct = (pl / st["cost"] * 100.0) if st["cost"] else None
        summary_data.append([
            a, str(st["n"]), fmt_money(st["cost"]), fmt_money(st["value"]),
            fmt_signed_money(st["loss"]), fmt_signed_money(st["profit"]),
            fmt_signed_money(pl), fmt_pct(pct),
        ])
    summary_data.append([
        "All", str(len(records)), fmt_money(tot_cost), fmt_money(tot_value),
        fmt_signed_money(tot_loss), fmt_signed_money(tot_profit),
        fmt_signed_money(tot_pl), fmt_pct(tot_pct),
    ])

    summary = Table(
        summary_data,
        colWidths=[0.75*inch, 0.55*inch, 1.3*inch, 1.3*inch, 1.3*inch,
                   1.3*inch, 1.3*inch, 0.9*inch],
        hAlign="LEFT",
    )
    summary_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f3864")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#b0b7c3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#1f3864")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2f9")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        # Total Losses always red, Total Profits always green.
        ("TEXTCOLOR", (4, 1), (4, -1), red),
        ("TEXTCOLOR", (5, 1), (5, -1), green),
    ])
    # Color each account's net P/L and % columns by sign.
    for r in range(1, len(summary_data)):
        plv = to_float(summary_data[r][6].replace("$", "").replace(",", "")
                       .replace("+", ""))
        summary_style.add("TEXTCOLOR", (6, r), (7, r),
                          green if (plv is None or plv >= 0) else red)
    summary.setStyle(summary_style)
    elements.append(summary)

    doc.build(elements)

    # ---- Console summary (grouped by account) ----
    print(f"Processed {len(records)} individual lots, grouped by account "
          f"({order_desc}).")
    for code, label in GROUP_LABELS:
        recs = [r for r in records if r["account"] == code]
        st = acct_stats[code]
        print(f"\n=== {label} (Account {code}) - {len(recs)} lots"
              f" | Total Loss: {fmt_signed_money(st['loss'])}"
              f" | Total Profit: {fmt_signed_money(st['profit'])} ===")
        print(f"{'Ticker':<8}{'Row':>5}{'Count':>12}{'CurPrice':>12}"
              f"{'TotCost':>14}{'CurValue':>14}{'P/L':>14}{'%Chg':>9}")
        for rec in recs:
            print(f"{rec['ticker']:<8}{rec['row']:>5}"
                  f"{fmt_qty(rec['count']):>12}"
                  f"{fmt_money(rec['current_price']):>12}"
                  f"{fmt_money(rec['total_cost']):>14}"
                  f"{fmt_money(rec['current_value']):>14}"
                  f"{fmt_signed_money(rec['profit_loss']):>14}"
                  f"{fmt_pct(rec['pct_change']):>9}")
        g_pl = st["value"] - st["cost"]
        g_pct = (g_pl / st["cost"] * 100.0) if st["cost"] else None
        print(f"{'SUBTOTAL':<25}{'':>12}{fmt_money(st['cost']):>14}"
              f"{fmt_money(st['value']):>14}{fmt_signed_money(g_pl):>14}"
              f"{fmt_pct(g_pct):>9}")

    print("\nProfit / Loss by account")
    print(f"{'Acct':<6}{'Lots':>6}{'Total Cost':>16}{'Current Value':>16}"
          f"{'Total Losses':>16}{'Total Profits':>16}{'P/L':>16}{'%Chg':>9}")
    for a in ACCOUNTS:
        st = acct_stats[a]
        pl = st["value"] - st["cost"]
        pct = (pl / st["cost"] * 100.0) if st["cost"] else None
        print(f"{a:<6}{st['n']:>6}{fmt_money(st['cost']):>16}"
              f"{fmt_money(st['value']):>16}{fmt_signed_money(st['loss']):>16}"
              f"{fmt_signed_money(st['profit']):>16}{fmt_signed_money(pl):>16}"
              f"{fmt_pct(pct):>9}")
    print(f"{'All':<6}{len(records):>6}{fmt_money(tot_cost):>16}"
          f"{fmt_money(tot_value):>16}{fmt_signed_money(tot_loss):>16}"
          f"{fmt_signed_money(tot_profit):>16}{fmt_signed_money(tot_pl):>16}"
          f"{fmt_pct(tot_pct):>9}")
    print(f"\nPDF written to: {out_path}")


if __name__ == "__main__":
    main()
